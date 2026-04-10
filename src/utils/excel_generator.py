import tempfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def generate_batch_report_excel(batch_data: dict, products: list[dict], stats: dict) -> str:
    """Генерирует Excel отчёт по партии. Возвращает путь к временному файлу."""
    wb = Workbook()

    # Лист 1: Информация о партии
    ws1 = wb.active
    ws1.title = "Информация о партии"
    header_font = Font(bold=True, size=12)
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 40

    info_rows = [
        ("Номер партии", batch_data.get("batch_number")),
        ("Дата партии", batch_data.get("batch_date")),
        ("Статус", "Закрыта" if batch_data.get("is_closed") else "Открыта"),
        ("Рабочий центр", batch_data.get("work_center_name", "")),
        ("Смена", batch_data.get("shift")),
        ("Бригада", batch_data.get("team")),
        ("Номенклатура", batch_data.get("nomenclature")),
        ("Код ЕКН", batch_data.get("ekn_code")),
        ("Начало смены", batch_data.get("shift_start")),
        ("Окончание смены", batch_data.get("shift_end")),
    ]
    for i, (label, value) in enumerate(info_rows, 1):
        ws1.cell(row=i, column=1, value=label).font = header_font
        ws1.cell(row=i, column=2, value=str(value) if value else "-")

    # Лист 2: Продукция
    ws2 = wb.create_sheet("Продукция")
    headers = ["ID", "Уникальный код", "Аггрегирована", "Дата аггрегации"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, product in enumerate(products, 2):
        ws2.cell(row=i, column=1, value=product.get("id"))
        ws2.cell(row=i, column=2, value=product.get("unique_code"))
        ws2.cell(row=i, column=3, value="Да" if product.get("is_aggregated") else "Нет")
        ws2.cell(row=i, column=4, value=str(product.get("aggregated_at", "-")) if product.get("aggregated_at") else "-")

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 25

    # Лист 3: Статистика
    ws3 = wb.create_sheet("Статистика")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20

    stat_rows = [
        ("Всего продукции", stats.get("total_products", 0)),
        ("Аггрегировано", stats.get("aggregated", 0)),
        ("Осталось", stats.get("remaining", 0)),
        ("Процент выполнения", f"{stats.get('aggregation_rate', 0)}%"),
    ]
    for i, (label, value) in enumerate(stat_rows, 1):
        ws3.cell(row=i, column=1, value=label).font = header_font
        ws3.cell(row=i, column=2, value=str(value))

    # Сохранение
    tmp = tempfile.NamedTemporaryFile(
        suffix=".xlsx",
        prefix=f"batch_{batch_data.get('batch_number', 'unknown')}_report_",
        delete=False,
    )
    wb.save(tmp.name)
    tmp.close()
    return tmp.name
